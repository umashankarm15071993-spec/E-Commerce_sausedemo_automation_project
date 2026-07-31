import openpyxl

class ExcelUtilities:

    @staticmethod
    def max_row(file,sheet):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet]
        return sheet.max_row

    @staticmethod
    def max_column(file,sheet):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet]
        return sheet.max_column

    @staticmethod

    def read_excel(file,sheet,n_row,n_clm):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet]
        return sheet.cell(n_row,n_clm).value

